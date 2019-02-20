# -*- coding: utf-8 -*-
import tempfile
import pandas as pd
import numpy as np
import re
import os
import sys
import win32com.client
import time
from tqdm import tqdm
import openpyxl
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
import msoffcrypto
import xlrd

# Set initial working path for the script
path = os.path.join('C:/Users/pphuc/Desktop/Docs/Current Using Docs/')


class BEC_Non_Domestic(object):
    def __init__(self, bec_file, sheetName, project_name, file_name, tab):
        self.fileName = file_name
        self.project_name = project_name
        self.sheetName = sheetName
        self.tab = tab
        self.sheet = pd.read_excel(bec_file, sheetName, keep_default_na=False, header=None).dropna(thresh=1)
        self.data_site_reference = ''
        self.data_site_measures = ''
        self.data_site_measure_unit = []

    # Extract data for site references
    def extract_site_reference(self, last_line_index):
        # Get the index (line number) of 'Project Category' as the beginning line of extracted data for site references
        project_category_index = self.sheet.iloc[:, 0][self.sheet.iloc[:, 0] == 'Project Category'].index.tolist()[0]
        # Extract data from 3 standalone columns
        small_df = pd.concat([self.sheet.loc[last_line_index - 3:last_line_index - 1, 3],
                              self.sheet.loc[last_line_index - 3:last_line_index - 1, 2]], axis=1, sort=False)
        # Re-organize these 3 standalone columns
        small_df = small_df.transpose().reset_index(drop=True).transpose()
        # Extract the rest of data
        TEMP_df = self.sheet.iloc[project_category_index:last_line_index, 0:2].reset_index(drop=True)
        # Remove empty fields
        list_empty = TEMP_df[TEMP_df[0] == ''].index.tolist()
        if (len(list_empty) > 0):
            TEMP_df = (TEMP_df.drop(list_empty, axis=0).reset_index(drop=True))
        # Return result after merging 2 sets of data
        return TEMP_df.append(small_df, ignore_index=True, sort=False)

    # Get data for the first half of site measures
    def extract_first_half_site_measures(self, begin_row_index, last_column_index,last_line):
        # Set a list of dropped columns of the first half of collected data
        list_dropped_columns_energy = ['Description of Minimum Data Required for Existing Specification',
                                       'Description of Minimum Data Required for Proposed Specification',
                                       'Additional Information']
        # Identify the index of columns needed to drop in the first half of data
        columns_to_drop_energy = self.sheet.iloc[begin_row_index + 1, 0:][
            self.sheet.iloc[begin_row_index + 1, 0:].isin(list_dropped_columns_energy)].index.tolist()
        # Extract the first half of data
        return self.sheet.iloc[begin_row_index + 1:last_line, 0:last_column_index].reset_index(drop=True).drop(
            columns_to_drop_energy, axis=1)

    # Get data for second half of site measures
    def extract_second_half_site_measures(self, begin_row_index, begin_column_index,last_line):
        # Identify the last column of second half data
        last_column_unit = self.sheet.iloc[begin_row_index, 0:][
            self.sheet.iloc[begin_row_index, 0:] == 'Energy Credits'].index.tolist()[-1]
        # Identify the index of columns needed to drop in the second half of data
        list_dropped_columns_unit = ['Milestone', 'Invoice', '', 'Milestone Claim', 'Amount']
        # Identify the index of columns needed to drop in the second half of data
        columns_to_drop_unit = self.sheet.iloc[begin_row_index, begin_column_index:last_column_unit + 1][
            self.sheet.iloc[begin_row_index, begin_column_index:last_column_unit + 1].isin(
                list_dropped_columns_unit)].index.tolist()
        # Extract the second half of data
        return self.sheet.iloc[begin_row_index:last_line, begin_column_index:last_column_unit + 1].drop(begin_row_index + 1,
                                                                                                 axis=0).reset_index(
            drop=True).drop(columns_to_drop_unit, axis=1)

    # Merge data of 2 halfs into 1 site measures
    def extract_site_measures(self, begin_row_index,last_line):
        # Identify the last column of first half set of collected data and starting column of the second half set of collected data
        col_index_energy_upgrades = self.sheet.iloc[begin_row_index, 0:][
            self.sheet.iloc[begin_row_index, 0:] == 'Electrical Savings kWh'].index.tolist()[0]
        TEMP_data_site_measures_proposed_energy_upgrades = self.extract_first_half_site_measures(begin_row_index,
                                                                                                 col_index_energy_upgrades,last_line)
        TEMP_data_site_measures_unit = self.extract_second_half_site_measures(begin_row_index,
                                                                              col_index_energy_upgrades,last_line)
        # Store the unit column for later checking
        self.data_site_measure_unit = TEMP_data_site_measures_unit.iloc[0].tolist()
        # Merge 2 sets of data
        TEMP_data_site_measures = pd.concat(
            [TEMP_data_site_measures_proposed_energy_upgrades, TEMP_data_site_measures_unit], axis=1, sort=False)
        # Reset the index column
        TEMP_data_site_measures.columns = [i for i in range(TEMP_data_site_measures.shape[1])]
        # Remove redundant texts
        return TEMP_data_site_measures.loc[~TEMP_data_site_measures[0].astype(str).isin(['Total', '', '-', ' '])]



    # Extract input data from non domestic tab
    def extract_data_from_input_sheet(self):
        # Identify the line distinguish between data for site reference and data for site measures
        try:
            proposed_engergy_upgrade_index = \
                self.sheet.iloc[:, 0][self.sheet.iloc[:, 0] == 'Proposed Energy Upgrades'].index.tolist()[0]
        except IndexError:
            proposed_engergy_upgrade_index = 14
        self.data_site_reference = self.extract_site_reference(proposed_engergy_upgrade_index)
        # Extract data for site measures
        last_line = self.sheet.iloc[:, 0][self.sheet.iloc[:, 0] == 'Total'].index.tolist()[0]
        self.data_site_measures = self.extract_site_measures(proposed_engergy_upgrade_index,last_line)
        return [self.data_site_measures, self.data_site_reference]


class BEC_project(object):
    def __init__(self, folder, file):
        self.file_name = file
        self.input_folder = path + folder + '/'
        self.out_put_folder = ''
        self.project_name = re.search(r'BEC(\s|\_?)\d+(\s?)\d+', file).group()
        self.project_year = re.search(r'\d+', self.input_folder).group()
        self.bec_file = pd.ExcelFile(self.input_folder + file)
        self.BEC_worksheet = {}
        self.empty_line = []
        self.beneficiary_dataframe = None
        self.project_summary_dataframe = None
        self.site_references = None
        self.site_measures = None
        self.site_measures_units = {}
        self.list_Add_addition_row_summary= None
        self.list_beginn_row_summary = None
        # Store each tab into dictionary
        for sheetName in self.bec_file.sheet_names:
            if ('Project Summary' == sheetName):
                self.BEC_worksheet[sheetName] = pd.read_excel(self.bec_file, sheetName, keep_default_na=False,
                                                              header=None)
            if ('Non Domestic' in sheetName):
                self.BEC_worksheet[sheetName] = BEC_Non_Domestic(self.bec_file, sheetName, self.project_name,
                                                                 self.file_name, sheetName)
            if ('Beneficiary' == sheetName):
                self.BEC_worksheet[sheetName] = pd.read_excel(self.bec_file, sheetName, keep_default_na=False,
                                                              header=None)

    # Prepare_section to get  data for summary data
    def prepare_section_limit_summary_data(self):
        # Convert the first column into all string type
        TEMP_dataframe2 = self.BEC_worksheet['Project Summary'].iloc[:, 0].astype(str)
        # Find beginning row index of the extracted data
        self.list_beginn_row_summary = TEMP_dataframe2[
            TEMP_dataframe2.str.contains('Better Energy Communities Programme - Non Domestic Costs',
                                         na=False)].index.tolist()
        # Find the last row index of the extracted data
        TEMP_dataframe = self.BEC_worksheet['Project Summary'].iloc[:, 1].astype(str)
        self.list_Add_addition_row_summary = TEMP_dataframe[TEMP_dataframe == 'Add additional rows as required'].index.tolist()
        if len(self.list_Add_addition_row_summary) == 0:
            self.list_Add_addition_row_summary = TEMP_dataframe2[
                TEMP_dataframe2.str.contains('Better Energy Communities Programme - Domestic Costs',
                                             na=False)].index.tolist()

    # Get data of the first half of requested table
    def first_half_summary_data(self, column_to_collect):
        TEMP_data_project_summary1 = self.BEC_worksheet['Project Summary'].iloc[
                                     self.list_beginn_row_summary[-1] + 2:self.list_Add_addition_row_summary[0],
                                     0:column_to_collect].reset_index(drop=True).drop(3, axis=1)
        # Get the empty value in the first half of table
        # list_0 = TEMP_data_project_summary1[(TEMP_data_project_summary1.loc[:,1] == 0) |(TEMP_data_project_summary1.iloc[1:,1]=='Facility Name')| (TEMP_data_project_summary1.iloc[1:,1]==' ')].index.tolist()
        list_0 = TEMP_data_project_summary1[(TEMP_data_project_summary1.loc[:, 1] == 0)].index.tolist()
        list_empty = TEMP_data_project_summary1[(TEMP_data_project_summary1[2] == '')].index.tolist()
        self.empty_line = list_0 + list_empty
        # Remove empty values
        TEMP_data_project_summary1 = TEMP_data_project_summary1.drop(self.empty_line, axis=0).reset_index(drop=True)
        if int(self.project_year) >= 2017:
            # Convert % num to numemric
            if (len(TEMP_data_project_summary1.iloc[:, 3].unique()) == 1 and
                    TEMP_data_project_summary1.iloc[:, 3].unique()[0] == u' '):
                TEMP_data_project_summary1.drop(4, axis=1, inplace=True)
            TEMP_data_project_summary1.iloc[1:, 3:] = TEMP_data_project_summary1.iloc[1:, 3:].fillna(0)
            TEMP_data_project_summary1.update((TEMP_data_project_summary1.iloc[1:, 3:] * 100).astype(float))
        return TEMP_data_project_summary1

    # Get data for the second half of requested data
    def second_half_summary_data(self):
        # Get index of the second half requested table
        header_line_boolean = self.BEC_worksheet['Project Summary'].iloc[self.list_beginn_row_summary[-1]].astype(str).isin(
            ['Total Project Cost', 'SEAI funding', 'Eligible VAT', 'SEAI Funding'])
        header_line_index = header_line_boolean[header_line_boolean == True].index.tolist()
        if int(self.project_year) == 2016:
            header_line_index = header_line_index[:3]
        # Extract data
        TEMP_data_project_summary2 = self.BEC_worksheet['Project Summary'].iloc[
                                     self.list_beginn_row_summary[-1]:self.list_Add_addition_row_summary[0], header_line_index].drop(
            [self.list_beginn_row_summary[-1] + 1, self.list_beginn_row_summary[-1] + 2], axis=0).reset_index(drop=True)
        TEMP_data_project_summary2 = TEMP_data_project_summary2.drop(self.empty_line, axis=0).reset_index(drop=True)
        return TEMP_data_project_summary2

    # Extract summary data
    def extract_summary_data(self):
        self.prepare_section_limit_summary_data()
        if (len(self.list_Add_addition_row_summary) == 1):
            if int(self.project_year) >= 2017:
                column_to_collect = 6
            else:
                column_to_collect = 4
            # GEt first half of summary data
            TEMP_data_project_summary1 = self.first_half_summary_data(column_to_collect)
            # Get data of the second half of requested table
            TEMP_data_project_summary2 = self.second_half_summary_data()
            # Merge 2 tables into 1
            data_project_summary = pd.concat([TEMP_data_project_summary1, TEMP_data_project_summary2], axis=1,
                                             sort=False)
            data_project_summary.insert(0, '-1', self.project_name)
            data_project_summary.iloc[0, 0] = 'Project Code'
            data_project_summary.iloc[0, 1] = 'Tab'
            data_project_summary.insert(0, '-2', self.project_year)
            data_project_summary.iloc[0, 0] = 'Year'
            self.project_summary_dataframe = data_project_summary
        else:
            print('Can not identify as there are more "Add additional rows as required" or no results in',self.file_name)

    # Extract beneficiary data in Beneficiary tab
    def extract_beneficiary_data(self):
        # Identify the first row index
        starting_scanning_line = self.BEC_worksheet['Beneficiary'].iloc[0:, 1][
            self.BEC_worksheet['Beneficiary'].iloc[:, 1] == 'Beneficiary Name'].index.tolist()[0]
        # Extract data
        TEMP_data_beneficiary = self.BEC_worksheet['Beneficiary'].iloc[starting_scanning_line:, 1]
        return TEMP_data_beneficiary

    # Extract beneficiary data in Summary tab
    def extract_beneficiary_data_in_summary(self):
        # Identify column contain Beneficiary
        series_contain_value = self.BEC_worksheet['Project Summary'].iloc[self.list_beginn_row_summary[0]+2]
        try:
            column_index = series_contain_value[series_contain_value.isin(['Beneficiary Name','Name of Beneficiary'])].index.tolist()[0]
        except Exception:
            print ('Cant find the index of column Beneficiary in Project Summary of ',self.file_name)
        # Extract data
        TEMP_data_beneficiary = self.BEC_worksheet['Project Summary'].iloc[self.list_beginn_row_summary[0]+2:self.list_Add_addition_row_summary[0],column_index].reset_index(drop=True).drop(self.empty_line, axis=0).reset_index(drop=True)
        return TEMP_data_beneficiary

    def finilize_beneficiary_extraction(self,TEMP_data_beneficiary):
        # Remove unnecessay data and reindex rows
        data_beneficiary = TEMP_data_beneficiary.loc[~TEMP_data_beneficiary.isin(
            ['Total Project Cost', '', 'Enter Name of Beneficiary', 'Enter Name of Beneficiary ', 'Name Of Beneficiary',0])].to_frame().reset_index(drop=True)
        # Add columns
        data_beneficiary.insert(0, 0, self.project_name)
        data_beneficiary.iloc[0, 0] = 'Project Code'
        data_beneficiary.insert(0, '-1', self.project_year)
        data_beneficiary.iloc[0, 0] = 'Year'
        data_beneficiary.iloc[0,-1]='Beneficiary Name'
        self.beneficiary_dataframe = data_beneficiary



    # Extract non domestic measures
    def extract_non_domestic_measure(self, non_domestic_sheet, list_measures):
        # Get measures
        non_domestic_measures = self.BEC_worksheet[non_domestic_sheet].extract_data_from_input_sheet()[0]
        # Add column
        non_domestic_measures.insert(0, '2', [i for i in range(non_domestic_measures.shape[0])])
        if len(list_measures) > 0:
            non_domestic_measures = non_domestic_measures.drop(0, axis=0)
        non_domestic_measures.insert(0, '1', non_domestic_sheet)
        # Add measure of each tab into a list
        list_measures.append(non_domestic_measures)
        # Store unit measure for later checking
        self.site_measures_units[non_domestic_sheet] = self.BEC_worksheet[non_domestic_sheet].data_site_measure_unit
        return list_measures

    # Check if all non domestic tabs in a BEC file having the same header format
    def check_site_measures_units_each_file(self):
        # check for an empty dictionary first if that's possible
        expected_value = list(self.site_measures_units.values())[0]
        all_equal = all(value == expected_value for value in self.site_measures_units.values())
        if all_equal:
            return True
        return False

    # Extract non domestic site reference
    def extract_non_domestic_reference(self, non_domestic_sheet, list_reference):
        # Get reference
        non_domestic_reference = self.BEC_worksheet[non_domestic_sheet].extract_data_from_input_sheet()[1].transpose()
        # Add column
        non_domestic_reference.insert(0, '2', int(re.search(r'\b\d+\b', non_domestic_sheet).group()))
        if len(list_reference) > 0:
            non_domestic_reference = non_domestic_reference.drop(0, axis=0)
        non_domestic_reference.insert(0, '1', non_domestic_sheet)
        # Add reference of each tab into a list
        list_reference.append(non_domestic_reference)
        return list_reference



    # Merge data of measures
    def merge_all_tabs_measures(self, list_measures):
        if (len(list_measures) > 0):
            # Merge data
            TEMP_site_measures_df = pd.concat(list_measures, ignore_index=True, sort=False)
            # Add column
            TEMP_site_measures_df.insert(0, '0', self.project_name)
            TEMP_site_measures_df.insert(0, '-1', self.project_year)
            TEMP_site_measures_df.iloc[0, 0] = 'Year'
            TEMP_site_measures_df.iloc[0, 1] = 'Project Code'
            TEMP_site_measures_df.iloc[0, 2] = 'Tab'
            TEMP_site_measures_df.iloc[0, 3] = 'ID Measures'
            self.site_measures = TEMP_site_measures_df


    # Merge data of references
    def merge_all_tabs_reference(self, list_reference):
        if (len(list_reference) > 0):
            # Merge
            TEMP_site_reference_df = pd.concat(list_reference, ignore_index=True, sort=False)
            # Add column
            TEMP_site_reference_df.insert(0, '0', self.project_name)
            TEMP_site_reference_df.insert(0, '-1', self.project_year)
            TEMP_site_reference_df.iloc[0, 0] = 'Year'
            TEMP_site_reference_df.iloc[0, 1] = 'Project Code'
            TEMP_site_reference_df.iloc[0, 2] = 'Tab'
            TEMP_site_reference_df.iloc[0, 3] = 'ID References'
            columns = TEMP_site_reference_df.iloc[0, :].reset_index(drop=True)
            floor_area = columns[columns == 'Floor Area of building'].index[0]
            TEMP_site_reference_df.insert(int(floor_area + 1), 'Unit', 'Unit')
            TEMP_site_reference_df.insert(int(floor_area + 1), 'Number', 'Num')
            # Split Floor Area of builing into 2 columns
            TEMP_site_reference_df.loc[1:, 'Unit'] = TEMP_site_reference_df.iloc[1:, int(floor_area)].astype(
                str).str.replace(r'\d+(\.?)\d+', '', regex=True)
            TEMP_site_reference_df.loc[1:, 'Number'] = \
                TEMP_site_reference_df.iloc[1:, int(floor_area)].astype(str).str.extract(r'(\d+(\.?)\d+)',
                                                                                         expand=False)[0]
            self.site_references = TEMP_site_reference_df


    #Tab needed to remove
    def list_remove_tab(self):
        dic_removed = {
            '2018': {
                'BEC00769': ['4','6','8','9','18'],
                'BEC00771': ['7','9'],
                'BEC00781': ['5','6','8','14','15','16'],
                'BEC00790': ['7','13'],
                'BEC00792': ['3'],
                'BEC00807': ['5','7'],
                'BEC00816': ['6','7','8','11','12'],
            },
            '2017':{
                'BEC 625': ['1','5','21','23'],
                'BEC 629': ['1','2'],
                'BEC 632': ['7'],
                'BEC 00633': ['7'],
                'BEC 647': ['1'],
                'BEC 648': ['6','8'],
                'BEC 653': ['3'],
                'BEC 661': ['7'],
                'BEC 662': ['1','2'],
                'BEC 668': ['5','16','17','19','28'],
                'BEC 672': ['1'],
                'BEC 673': ['18'],
                'BEC 675': ['6'],
                'BEC 678': ['3','4'],
                'BEC 679': ['4','5','6','7'],
                'BEC 710': ['6'],
                'BEC 711': ['1','3','5','6','7','8'],
                'BEC 718': ['2']
            },
            '2016': {
                'BEC 00 498': ['13'],
                'BEC 00466': ['2'],
                'BEC 00481': ['8','10','17'],
                'BEC 00485': ['5','6','8','9','13'],
                'BEC 00510': ['1','9'],
                'BEC 00517': ['3','13'],
                'BEC 00521': ['1'],
                'BEC 00522': ['1'],
                'BEC 00531': ['3'],
                'BEC 00532': ['3','4','5','6'],
                'BEC 00539': ['4'],
                'BEC 00540': ['4','6'],
                'BEC 00544': ['1','4','10'],
                'BEC 00563': ['2','5'],
                'BEC 00565': ['1'],
                'BEC 00575': ['3','5'],
                'BEC 00577': ['4','5','6']
            },
            '2015': {}
        }
        return dic_removed[self.project_year]


    # Collect data from each non domestic data tab in each project
    def extract_non_domestic_data(self):
        # List all non domestic tabs
        non_domestic_list = [i for i in self.BEC_worksheet.keys() if
                             'Non Domestic' in i and int(re.search(r'\b\d+\b', i).group()) in
                             self.project_summary_dataframe[0].tolist() and (self.project_name not in list(self.list_remove_tab().keys()) or re.search(r'\b\d+\b', i).group() not in self.list_remove_tab()[self.project_name])]
        list_measures = []
        list_reference = []
        # Iterating through each tab
        for non_domestic_sheet in non_domestic_list:
            # Non Domestic Measures
            list_measures = self.extract_non_domestic_measure(non_domestic_sheet, list_measures)
            # Non Domestic Reference
            list_reference = self.extract_non_domestic_reference(non_domestic_sheet, list_reference)
        # Non Domestic Measures
        self.merge_all_tabs_measures(list_measures)
        # Non Domestic Reference
        self.merge_all_tabs_reference(list_reference)

    # Function that controls extracting functions
    def extract_data(self):
        self.extract_summary_data()
        # If Beneficiary tab in list of tabs then extract or else look at Project Summary tab
        if 'Beneficiary' in self.bec_file.sheet_names:
            temp_beneficiary = self.extract_beneficiary_data()
            self.finilize_beneficiary_extraction(temp_beneficiary)
        else:
            temp_beneficiary = self.extract_beneficiary_data_in_summary()
            self.finilize_beneficiary_extraction(temp_beneficiary)
        self.extract_non_domestic_data()

    # Checking if attributes are available or not
    def check_available_result(self):
        if (self.project_summary_dataframe is not None and self.site_references is not None and
                self.site_measures.shape[0] is not None):
            return True
        else:
            return False

    # Write individual project into seperate files
    def write_seperate_excel_file(self, folder_name):
        # Create a shared folder along side with year
        if not os.path.exists(path + folder_name + ' Shared Data/'):
            os.makedirs(path + folder_name + ' Shared Data/')
        new_path = path + folder_name + ' Shared Data/'
        # Create a folder for a project
        if not os.path.exists(new_path + self.project_name + '/'):
            os.makedirs(new_path + self.project_name + '/')
        new_path += self.project_name + '/'
        self.out_put_folder = new_path
        # Store data and write them to excel files
        if (self.project_summary_dataframe is not None):
            self.project_summary_dataframe.to_excel(self.out_put_folder + self.project_name + '_Project Summary.xlsx',
                                                    'Project Summary', header=False, index=False)
        if (self.beneficiary_dataframe is not None):
            self.beneficiary_dataframe.to_excel(self.out_put_folder + self.project_name + '_Beneficiary.xlsx',
                                                'Beneficiary', header=False, index=False)
        if (self.site_references is not None and self.site_measures is not None):
            self.site_references.to_excel(self.out_put_folder + self.project_name + '_References.xlsx', 'References',
                                          header=False, index=False)
            self.site_references.to_excel(self.out_put_folder + self.project_name + '_References.xlsx', 'References',
                                          header=False, index=False)
            self.site_measures.to_excel(self.out_put_folder + self.project_name + '_Measures.xlsx', 'Measures',
                                        header=False, index=False)

    # Checking for missing headers in both dataframe
    def checking_missing_headers(self, current_df, extracted_df, file_name):
        # Check how different they are for current headers with extracted headers from excel
        if check_different(current_df.iloc[0, :].astype(str).tolist(), extracted_df.iloc[0, :].astype(str).tolist()):
            # Check for missing column in extracted headers
            extracted_df_index_missing = find_difference(current_df.iloc[0, :].astype(str).tolist(),
                                                         extracted_df.iloc[0, :].astype(str).tolist(), 'missing')
            # If missing column then add empty column
            if extracted_df_index_missing is not None and len(extracted_df_index_missing) > 0:
                # Extract again the whole dataframe
                extracted_df = pd.read_excel(self.out_put_folder + file_name + '.xlsx', file_name,
                                             keep_default_na=False, header=None, index=False)
                for new_column in extracted_df_index_missing:
                    # new_column[1] += extracted_df_index_missing.index(new_column)
                    extracted_df = fill_empty_value_into_blank_columns(new_column, extracted_df)
                extracted_df.to_excel(self.out_put_folder + file_name + '.xlsx', file_name, header=False, index=False)
        # Check how different they are between current headers and extracted headers
        if check_different(extracted_df.iloc[0, :].astype(str).tolist(), current_df.iloc[0, :].astype(str).tolist()):
            # Check for missing column in current headers
            current_df_index_missing = find_difference(extracted_df.iloc[0, :].astype(str).tolist(),
                                                       current_df.iloc[0, :].astype(str).tolist(), 'missing')
            # Add column if missing
            if current_df_index_missing is not None and len(current_df_index_missing) > 0:
                for new_column in current_df_index_missing:
                    # new_column[1] += current_df_index_missing.index(new_column)
                    current_df = fill_empty_value_into_blank_columns(new_column, current_df)
        return current_df, extracted_df

    # Checking for different headers in both dataframe
    def checking_different_headers(self, current_df, extracted_df):
        # Check how different they are for current headers with extracted headers from excel
        if check_different(current_df.iloc[0, :].astype(str).tolist(),
                           extracted_df.iloc[0, :].astype(str).tolist()):
            # Check for different column in extracted headers
            extracted_df_index_different = find_difference(current_df.iloc[0, :].astype(str).tolist(),
                                                           extracted_df.iloc[0, :].astype(str).tolist(),
                                                           'different')
            # If different then replace with the latest change
            if extracted_df_index_different is not None and len(extracted_df_index_different) > 0:
                if (int(self.project_year) > 2018):
                    for column in extracted_df_index_different:
                        extracted_df.iloc[0, column[1]] = column[0]
                else:
                    for column in extracted_df_index_different:
                        current_df.iloc[0, column[1]] = extracted_df.iloc[0, column[1]]
        # Check how different they are for current headers with extracted headers from excel
        if check_different(extracted_df.iloc[0, :].astype(str).tolist(),
                           current_df.iloc[0, :].astype(str).tolist()):
            # Check for different column in current headers
            current_df_index_different = find_difference(extracted_df.iloc[0, :].astype(str).tolist(),
                                                         current_df.iloc[0, :].astype(str).tolist(), 'different')
            # If different then replace with the latest change
            if current_df_index_different is not None and len(current_df_index_different) > 0:
                if (int(self.project_year) > 2018):
                    for column in current_df_index_different:
                        extracted_df.iloc[0, column[1]] = column[0]
                else:
                    for column in current_df_index_different:
                        current_df.iloc[0, column[1]] = extracted_df.iloc[0, column[1]]
        return current_df, extracted_df

    # Add data into an excel file
    def add_project(self):
        # Used these lines for writting to seperated folders for each year
        # if not os.path.exists(path + 'BEC ' + self.project_year + ' Shared Data/'):
        #     os.makedirs(path + 'BEC ' + self.project_year + ' Shared Data/')
        # self.out_put_folder = path + 'BEC ' + self.project_year + ' Shared Data/'
        # Create a shared folder for all files in all years
        if not os.path.exists(path + 'BEC Shared Data/'):
            os.makedirs(path + 'BEC Shared Data/')
        self.out_put_folder = path + 'BEC Shared Data/'
        #Write each tabs into seperate files
        #self.write_files(self.project_summary_dataframe, 'Project Summary')
        if (self.beneficiary_dataframe is not None):
            self.write_files(self.beneficiary_dataframe, 'Beneficiary')
        #self.write_files(self.site_measures, 'Site Measures')
        #self.write_files(self.site_references, 'Site References')

    # Write all data into 4 shared files
    def write_files(self, dataframe, file_name):
        # Create a shared file
        if not (os.path.isfile(self.out_put_folder + file_name + '.xlsx')):
            dataframe.to_excel(self.out_put_folder + file_name + '.xlsx', file_name, header=False, index=False)
        else:
            # Initialize data for validation
            dataframe.columns = [i for i in range(len(dataframe.columns))]
            current_df = dataframe
            extracted_df = pd.read_excel(self.out_put_folder + file_name + '.xlsx', file_name, keep_default_na=False,
                                         header=None, index=False, nrows=1)
            # To see what makes different
            if current_df.iloc[0, :].astype(str).tolist() != extracted_df.iloc[0, :].astype(str).tolist():
                # Checking for missing headers in both dataframe
                current_df, extracted_df = self.checking_missing_headers(current_df, extracted_df, file_name)
                # Checking for different headers in both dataframe
                current_df, extracted_df = self.checking_different_headers(current_df, extracted_df)
            # Write to output files
            if current_df.iloc[0, :].astype(str).tolist() == extracted_df.iloc[0, :].astype(str).tolist():
                book = load_workbook(self.out_put_folder + file_name + '.xlsx')
                writer = pd.ExcelWriter(self.out_put_folder + file_name + '.xlsx', engine='openpyxl')
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                current_df.iloc[1:, :].to_excel(writer, file_name, index=False, header=False,
                                                startrow=writer.sheets[file_name].max_row)
                writer.save()
            else:
                print(self.project_name, 'hasnt printed', file_name,
                      'probably because of mismatch headers between files (not tabs)')


# Check 2 lists if they are different or not
def check_different(list1, list2):
    if len(list(set(list1) - set(list2))) > 0:
        return True
    return False


# Checking a text with a list of text to see how different it is. If it is completely different then it is marked as "missing" or "different" if not so much different
def check_header(text, list_text):
    for i in list_text:
        if fuzz.ratio(i, text) >= 92:
            return ['different', list_text.index(i)]
    return ['missing']


# Filling empty values into blank columns of a dataframe that doesnt have the column as in another dataframe
def fill_empty_value_into_blank_columns(new_column, current_df):
    current_df.insert(new_column[1], 'Empty Value', new_column[0])
    current_df.columns = [i for i in range(len(current_df.columns))]
    current_df.iloc[1:, new_column[1]] = ''
    return current_df


# Find the differences between 2 lists. A flag is passed into the function to ask for the corresponding results
def find_difference(list1, list2, flag):
    # Initialize values
    index_list_different, index_list_missing = None, None
    if len(list(set(list1) - set(list2))):
        # A list contains different between list 1 with list 2 => what list2 doesnt have
        diff = list(set(list1) - set(list2))
        # Check how different these elements. If completely different => Missing or else Different
        index_list_different = [[i, list1.index(i), check_header(i, list2)[1]] for i in diff if
                                check_header(i, list2)[0] == 'different']
        index_list_missing = [[i, list1.index(i)] for i in diff if check_header(i, list2)[0] == 'missing']
        # Sort the lsit
        index_list_different.sort(key=lambda x: x[1])
        index_list_missing.sort(key=lambda x: x[1])
    if (flag == 'missing'):
        return index_list_missing
    if (flag == 'different'):
        return index_list_different
    return


# Unprotect files if necessary
def unprotect_xlsm_file(path, filename, passw):
    xcl = win32com.client.Dispatch('Excel.Application')
    # Pass for files in 2018 'Bec2018dec2017'
    # Pass for files in 2017 'Bec141116'
    # Pass for files in 2016 'bec060314'
    # Pass for files in 2015 'bec050314'
    pw_str = passw
    wb = xcl.Workbooks.Open(path + filename, False, True, None, pw_str)
    xcl.DisplayAlerts = False
    wb.SaveAs(path + filename + 'x', None, '', '')
    xcl.Quit()


# List all files in a folder
def access_to_working_file(folder_name):
    files = os.listdir(path + folder_name)
    return files


# Executing each project file in a year
def execute_each_project_in_a_year(folder_name):
    file_list = access_to_working_file(folder_name)
    errors = []
    if (len(file_list) > 0):
        for file_name in tqdm(file_list):
            if ('.xlsm' in file_name or '.xlsx' in file_name or '.xls' in file_name):
                try:
                    temp_file = BEC_project(folder_name, file_name)
                    temp_file.extract_data()
                    if (temp_file.check_site_measures_units_each_file() == False):
                        print('There is an error with header in',temp_file.file_name)
                    else:
                        if (temp_file.check_available_result()):
                            # temp_file.write_seperate_excel_file(folder_name)
                            temp_file.add_project()
                except Exception:
                    errors.append(temp_file.project_name + ' from ' + temp_file.file_name)
    else:
        print('Folder ' + folder_name + ' is empty')
    if (len(errors) > 0):
        print('')
        print('Errors: ', len(errors), errors)


# Get the path for the working folder (BEC [year]), for example: BEC 2018 or BEC 2017
def working_with_folder():
    folder_list = os.listdir(path)
    for folder_name in folder_list[::-1]:
        if re.search(r'^BEC \d+$', folder_name):
            print('Executing folder', folder_name)
            execute_each_project_in_a_year(folder_name)


# Extract data randomly
def extract_randomly_data():
    selected_folder = input('Choose folder to select: ')
    selected_file = input('Choose file to select: ')
    extracted_headers = pd.read_excel(path + selected_folder + ' Shared Data/' + selected_file + '.xlsx', selected_file,
                                      keep_default_na=False, header=None, index=False, nrows=1)
    extracted_data = pd.read_excel(path + selected_folder + ' Shared Data/' + selected_file + '.xlsx', selected_file,
                                   keep_default_na=False, header=None, index=False)
    numb_of_rand_data = int(input('Number of data points: '))
    random_selected_data = extracted_data.iloc[1:, :].sample(numb_of_rand_data)
    result = extracted_headers.append(random_selected_data, sort=False)
    result.to_excel(selected_folder + ' ' + selected_file + ' ' + str(numb_of_rand_data) + ' searching results.xlsx',
                    header=False, index=False)
    print('Done!')


def main():
    # option = input('Choose your task (1 for executing files or 2 for randomly selecting data points): ')
    option = '1'
    if (option == '1'):
        start_time = time.time()
        working_with_folder()
        print('Done! from ', time.asctime(time.localtime(start_time)), ' to ',
              time.asctime(time.localtime(time.time())))


if __name__ == '__main__':
    main()


