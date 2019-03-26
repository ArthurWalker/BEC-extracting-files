# -*- coding: utf-8 -*-
import tempfile
import pandas as pd
import numpy as np
import re
import os
import sys
import win32com.client
import time
from tqdm import tqdm
import xlwings as xw
import openpyxl
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
import msoffcrypto
import xlrd

# Initilize working folder
path = os.path.join('C:/Users/pphuc/Desktop/Docs/Current Using Docs/')

# Write dataframe into excel files
def write_file(path,df,new_file_name):
    # List all row indexies whose first column is empty
    empty_list = df[df.iloc[:,1]==''].index.tolist()
    if (len(empty_list) > 0):
        # Remove rows in df from the list
        df = (df.drop(empty_list, axis=0).reset_index(drop=True))
    # Create a shared folder along side with year
    new_path = path +'Shared Data/'
    # If the path contains Evaluation then the path will be shorten to store result in the same folder as others
    if re.search('Evaluations',new_path):
        new_path = re.sub(r'BEW 2012/Evaluations/','',new_path)
    # If the folder doesnt exist then create one and create a file that has the first dataframe
    if not os.path.exists(new_path):
        os.makedirs(new_path)
        if not (os.path.isfile(new_path + new_file_name+'.xlsx')):
            df.to_excel(new_path + new_file_name+'.xlsx',new_file_name,header=False, index=False)
    else:
        # If the folder exists but not the file then create the file to store the first dataframe
        if not (os.path.isfile(new_path + new_file_name+'.xlsx')):
            df.to_excel(new_path + new_file_name+'.xlsx',new_file_name,header=False, index=False)
        else:
        # If the folder exists as well as the file then append to the file
            # Load workbook
            book = load_workbook(new_path +new_file_name+'.xlsx')
            # Get excel file
            writer = pd.ExcelWriter(new_path +new_file_name+'.xlsx',engine='openpyxl')
            writer.book = book
            # List all sheet names in the excel file
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            # Append dataframe to excel files while removing the first line as the first line is the headings which are already exist after the first dataframe is written
            df.iloc[1:, :].to_excel(writer, new_file_name, index=False, header=False,startrow=writer.sheets[new_file_name].max_row)
            # Save the excel file
            writer.save()

# Extract data from spread sheets
def extract_data(excel_file,tab,extracted_lst,skiprow,project_year):
    # Import data from excel to dataframe
    temp_df = pd.read_excel(excel_file, tab, skiprows=skiprow,keep_default_na=False, header=None)
    # Take the first line
    df_1_line = temp_df.iloc[0]
    if tab == 'Technologies':
        # Get the second column to the end with all rows
        extracted_df = temp_df.iloc[:,2:]
        # Add year column
        extracted_df.insert(0, '',project_year)
        extracted_df.iloc[0, 0] = 'Year'
    else:
        # Find indexies of all columns in the list
        col_index_workbook = find_column(df_1_line, extracted_lst)
        # Extend some columns
        col_extended_workbook = find_extended_column(tab, df_1_line, col_index_workbook)
        # Get all columns in the list
        extracted_df = temp_df[col_extended_workbook]
        # Add year column
        extracted_df.insert(0, '',project_year)
        extracted_df.iloc[0, 0] = 'Year'
    return extracted_df

# There are some columns need to be extracted to extend
def find_extended_column(tab,df_l_line,num_list):
    # Sort the index list
    num_list = sorted(num_list)
    new_num_lst = []
    if tab == 'BE Workplaces main workbook':
        # Get indexies of columns
        start_point = df_l_line[df_l_line=='Select Thermal Fuel'].index.tolist()[0]
        end_point = df_l_line[df_l_line == 'Total Energy Cost Savings €'].index.tolist()[0]
        start_point2 = df_l_line[df_l_line=='Primary Energy Savings kWh'].index.tolist()[0]
        end_point2 = df_l_line[df_l_line == 'Site Energy Reduction %'].index.tolist()[0]
        num_list[num_list.index(end_point2)] = end_point2-1
        end_point2 = end_point2-1
        # Extend all columns
        new_num_lst = num_list[:num_list.index(start_point)]+list(range(start_point,end_point+1))+num_list[num_list.index(end_point)+1:num_list.index(start_point2)+1]+list(range(start_point2,end_point2))+num_list[num_list.index(end_point2):]
    if len(new_num_lst)==0:
        new_num_lst = num_list
    return new_num_lst

# Find all indexies of columns which are in the list
def find_column(df_1_line,lst_to_find):
    return df_1_line[df_1_line.astype(str).isin(lst_to_find)].index.tolist()

# Working with Evaluation files
def assign_task_Evaluation(seeep_path,folder):
    input_folder = seeep_path + folder + '/'
    project_year = '2012'
    # List files in the folder
    file_path_lst = os.listdir(input_folder)
    # Iterate through each file in the folder
    for file in tqdm(file_path_lst):
        if re.search(r'Batch',file):
            # Get the excel file
            excel_file = pd.ExcelFile(input_folder+file)
            col_lst = ['Reference', 'Applicant', 'Description']
            # Extract column in dataframe from the list
            summary_df = extract_data(excel_file,'Summary Sheet',col_lst,0,project_year)
            # Write result to excel file
            write_file(input_folder,summary_df,'Summary')

# Working with Summary file
def assign_task_Summary(seeep_path,file,folder):
    input_folder = seeep_path + folder + '/'
    project_year = re.search(r'\d+', folder).group()
    # Get the excel file
    excel_file = pd.ExcelFile(input_folder + file)
    if 'Admin' in excel_file.sheet_names:
        lst_col_admin = ['Reference No.','Cat. ','Submitted By','Project Title','County','Approved Funding']
        # Extract column in dataframe from the list
        admin_df = extract_data(excel_file,'Admin',lst_col_admin,1,project_year)
        # Write result to excel file
        write_file(seeep_path,admin_df,'Admin')

# Working with Overview file
def assign_task_Overview(seeep_path,file,folder):
    input_folder = seeep_path + folder + '/'
    project_year = re.search(r'\d+', folder).group()
    # Get excel file
    excel_file = pd.ExcelFile(input_folder + file)
    if 'BE Workplaces main workbook' in excel_file.sheet_names:
        # List column needed to be extracted
        lst_col_workbook = ['SEAI Reference', 'Organisation', 'Project Title', 'Total Incl VAT', 'Total Excl VAT',
                            'Select Thermal Fuel', 'Total Energy Cost Savings €', 'Grant  /Approved (Proposed)',
                            'Grant /Approved (Proposed)', 'Primary Energy Savings kWh', 'Site Energy Reduction %']
        # Extract column in dataframe from the list
        df_workplaces = extract_data(excel_file,'BE Workplaces main workbook',lst_col_workbook,3,project_year)
        # Write result to excel file
        write_file(seeep_path, df_workplaces,'Workplaces')
    if 'Technologies' in excel_file.sheet_names:
        lst_col_tech = []
        tech_df = extract_data(excel_file,'Technologies',lst_col_tech,0,project_year)
        write_file(seeep_path, tech_df,'Technologies')

# Assign task for each file
def execute_each_folder(seeep_path,folder_name):
    file_path = seeep_path+folder_name+'/'
    file_path_lst = os.listdir(file_path)
    for file in tqdm(file_path_lst):
        if file == 'Evaluations':
            assign_task_Evaluation(seeep_path+'BEW 2012/',file)
        if re.search(r'Better Energy',file) and re.search(r'Summary',file):
            assign_task_Summary(seeep_path,file,folder_name)
        if re.search(r'Better Energy Board',file):
            assign_task_Overview(seeep_path,file,folder_name)


def main():
    #start_time = time.time()
    path_lst = os.listdir(path)
    if 'SEEEP' in path_lst:
        seeep_path = path+'SEEEP/'
        folder = os.listdir(seeep_path)
        for folder_name in folder:
            if re.search(r'BEW',folder_name):
                execute_each_folder(seeep_path,folder_name)
    #print('Done! from ', time.asctime(time.localtime(start_time)), ' to ',time.asctime(time.localtime(time.time())))

if __name__ == '__main__':
    main()
